"""Build the 8-tab MLB K-Prop workbook — portable version.

Run with: python build_workbook.py 2026-07-09 --slate am

Styling/layout logic below is UNCHANGED from the original. Only the
input/output paths were changed to route through WORKSPACE and be
parameterized by date instead of hardcoded to one day.
"""
import argparse
import json
import os
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from _paths import WORKSPACE

ap = argparse.ArgumentParser(description="Build the 8-tab K-prop workbook for a given slate")
ap.add_argument("date", nargs="?", default=datetime.now().strftime("%Y-%m-%d"))
ap.add_argument("--slate", choices=["am", "late"], default="am")
_args = ap.parse_args()
DATE_STR = _args.date
SLATE_TYPE = _args.slate

# Style constants
NAVY = "1A2744"
AMBER = "FFC000"
RED_FILL = "F8CBAD"
GREEN_FILL = "C6E0B4"
YELLOW_FILL = "FFE699"
PINK_FILL = "F4B6C2"
GRAY_FILL = "D9D9D9"
LIGHT_GRAY = "F2F2F2"

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

results = json.load(open(WORKSPACE / f"model_results_{DATE_STR}_{SLATE_TYPE}.json"))
_slate_raw = json.load(open(WORKSPACE / f"slate_{DATE_STR}_{SLATE_TYPE}.json"))
slate = {"games": _slate_raw} if isinstance(_slate_raw, list) else _slate_raw

wb = Workbook()
wb.remove(wb.active)


def style_header(ws, row, columns):
    """Apply navy header styling to a row."""
    for col_idx, hdr in enumerate(columns, start=2):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


def auto_size(ws, columns, start_col=2):
    """Set column widths based on header length + heuristic."""
    for i, h in enumerate(columns):
        col_letter = get_column_letter(start_col + i)
        ws.column_dimensions[col_letter].width = max(11, min(28, len(str(h)) + 4))


def fmt_cell(cell, val, num_fmt=None, fill=None, font=None, alignment=None, border=THIN_BORDER):
    cell.value = val
    cell.border = border
    if num_fmt:
        cell.number_format = num_fmt
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment


def safe_num(v):
    return v if isinstance(v, (int, float)) else None


# ── TAB 1: Summary ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Summary")
ws.column_dimensions["A"].width = 3
ws["B2"] = "MLB K-Prop Model — Sunday, July 5, 2026"
ws["B2"].font = Font(name="Calibri", bold=True, size=18, color=NAVY)
ws.merge_cells("B2:M2")
ws["B3"] = "v3.5 production + v3.6.1 + v3.5p SHADOW (gates still on v3.5) | Park factors v3.6.2 (Z-Files 2026) | csv-47 cohort (411)"
ws["B3"].font = Font(name="Calibri", italic=True, size=10, color="595959")
ws.merge_cells("B3:M3")

# Slate overview
ws["B5"] = "Slate Overview"
ws["B5"].font = Font(name="Calibri", bold=True, size=13, color=NAVY)
overview = [
    ("Games on slate", len(slate["games"])),
    ("Starting pitchers", len(results)),
    ("In FG cohort", sum(1 for r in results if r.get("FG Source") and r["FG Source"] != "missing")),
    ("BF Picks (v3.5 gate)", sum(1 for r in results if r.get("BF Call") == "Under")),
    ("No call", sum(1 for r in results if r.get("BF Call") == "No call")),
    ("Watch List", sum(1 for r in results if r.get("BF Call") == "Watch List")),
    ("FG CSV", "fangraphs-leaderboards-47.csv (411 rows)"),
    ("Park factors", "v3.6.2 / Z-Files 2026 (as of 2026-05-29)"),
    ("v4.0 status", "SHADOW — projections logged but BF gates still gated on v3.5 residuals"),
]
for i, (label, val) in enumerate(overview, start=6):
    ws.cell(row=i, column=2, value=label).font = Font(name="Calibri", bold=True)
    ws.cell(row=i, column=3, value=val)
    ws.cell(row=i, column=2).alignment = LEFT
    ws.cell(row=i, column=3).alignment = LEFT
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 50

# v4.0 shadow table
summary_start_row = 17
ws.cell(row=summary_start_row, column=2, value="v3.5 vs v4.0 Shadow Snapshot (today's slate)").font = Font(name="Calibri", bold=True, size=13, color=NAVY)
ws.cell(row=summary_start_row, column=2).alignment = LEFT

summary_cols = ["Pitcher", "Matchup", "K/9", "v3.5 Proj", "v4.0 Proj", "Δ (v4-v3.5)", "Best Line", "v3.5 Resid", "v4.0 Resid", "BF Call"]
style_header(ws, summary_start_row + 1, summary_cols)
ws.freeze_panes = "A20"  # row 3 frozen for navigation but summary table has its own header at 18

r = summary_start_row + 2
for row in results:
    if row.get("Proj Ks") is None and row.get("Proj Ks v4.0") is None:
        continue
    delta = row.get("Delta v4-v35")
    delta_fill = None
    if isinstance(delta, (int, float)):
        if delta > 0.05:
            delta_fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
        elif delta < -0.05:
            delta_fill = PatternFill(start_color=RED_FILL, end_color=RED_FILL, fill_type="solid")

    bf_fill = None
    if row.get("BF Call") == "Under":
        bf_fill = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")

    vals = [
        (row["name"], None, None),
        (row.get("matchup", ""), None, None),
        (row.get("K/9"), "0.00", None),
        (row.get("Proj Ks"), "0.00", None),
        (row.get("Proj Ks v4.0"), "0.00", None),
        (row.get("Delta v4-v35"), "+0.00;-0.00;0.00", delta_fill),
        (row.get("Best Line"), "0.0", None),
        (row.get("Residual"), "+0.00;-0.00;0.00", None),
        (row.get("Residual v4.0"), "+0.00;-0.00;0.00", None),
        (row.get("BF Call"), None, bf_fill),
    ]
    for ci, (v, fmt, fill) in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.border = THIN_BORDER
        cell.alignment = CENTER if ci > 2 else LEFT
        if fmt:
            cell.number_format = fmt
        if fill:
            cell.fill = fill
    r += 1

# Column widths
for ci, w in enumerate([24, 14, 8, 11, 11, 13, 11, 12, 12, 12]):
    ws.column_dimensions[get_column_letter(2 + ci)].width = w

# Footer note on v4.0
ws.cell(row=r + 1, column=2, value="v4.0 NOTE").font = Font(name="Calibri", bold=True, color=NAVY)
ws.cell(row=r + 2, column=2,
        value="v4.0 projections use today's FG export (live cohort). Backtest reference numbers "
              "(MAE 1.69, R² 0.23, DirAcc 64.4%) were scored against a 5/29 FG snapshot. "
              "Shadow comparison measures v4.0 vs v3.5 on identical inputs; both use today's FG export, "
              "so the relative delta is methodologically clean. Flags MATCH backtest: book_line=None, "
              "apply_short_mixture=False, apply_calibration=False. BF picks STILL gated on v3.5 residuals.")
ws.cell(row=r + 2, column=2).alignment = LEFT
ws.cell(row=r + 2, column=2).font = Font(name="Calibri", italic=True, size=9, color="595959")
ws.merge_cells(start_row=r + 2, start_column=2, end_row=r + 2, end_column=11)
ws.row_dimensions[r + 2].height = 45

# Apply freeze pane (row 3 per spec)
ws.freeze_panes = "A4"


# ── TAB 2: Raw Pitcher Stats ───────────────────────────────────────────────
ws = wb.create_sheet("Raw Pitcher Stats")
ws.column_dimensions["A"].width = 3
ws["B2"] = "Raw Pitcher Statistics — FanGraphs + Baseball Savant"
ws["B2"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
ws.merge_cells("B2:R2")

raw_cols = ["Pitcher", "Team", "Hand", "Matchup", "GS", "IP", "K/9", "K%", "BB%",
            "K-BB%", "SwStr%", "CSW%", "Whiff%", "Hard-Hit%", "ERA", "FIP", "xERA", "FG Source"]
style_header(ws, 3, raw_cols)
ws.freeze_panes = "A4"

r = 4
for row in results:
    is_watch = row.get("BF Call") == "Watch List"
    row_fill = PatternFill(start_color=GRAY_FILL, end_color=GRAY_FILL, fill_type="solid") if is_watch else None
    vals = [
        row["name"], row.get("team", ""), row.get("hand", ""), row.get("matchup", ""),
        row.get("GS"), row.get("IP"), row.get("K/9"), row.get("K%"), row.get("BB%"),
        row.get("K-BB%"), row.get("SwStr%"), row.get("CSW%"), row.get("Whiff%"),
        row.get("Hard-Hit%"), row.get("ERA"), row.get("FIP"), row.get("xERA"),
        row.get("FG Source", ""),
    ]
    for ci, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.border = THIN_BORDER
        cell.alignment = CENTER if ci > 2 else LEFT
        if ci > 5 and isinstance(v, (int, float)):
            cell.number_format = "0.00"
        if row_fill:
            cell.fill = row_fill
    r += 1

for ci, w in enumerate([22, 7, 7, 14, 6, 7, 7, 7, 7, 8, 8, 8, 8, 9, 7, 7, 7, 12]):
    ws.column_dimensions[get_column_letter(2 + ci)].width = w


# ── TAB 3: Classification & Multipliers ────────────────────────────────────
ws = wb.create_sheet("Classification & Multipliers")
ws.column_dimensions["A"].width = 3
ws["B2"] = "Classification & Multipliers"
ws["B2"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
ws.merge_cells("B2:H2")

# Multiplier table
ws.cell(row=3, column=2, value="Class").font = Font(bold=True)
ws.cell(row=3, column=3, value="Multiplier").font = Font(bold=True)
ws.cell(row=3, column=4, value="Notes").font = Font(bold=True)
mults = [("SwStr-Dominant", 0.95, "SwStr%≥14"),
         ("Above-Zone", 0.93, "High K-rate, fly-ball-leaning"),
         ("Below-Zone", 0.92, "Ground-ball with K upside"),
         ("Mixed", 0.91, "DEFAULT; auto if SwStr%≥11 or Whiff%≥27"),
         ("CS-Dependent", 0.90, "Called-strike heavy")]
for i, (c, m, note) in enumerate(mults):
    ws.cell(row=4 + i, column=2, value=c)
    ws.cell(row=4 + i, column=3, value=m).number_format = "0.00"
    ws.cell(row=4 + i, column=4, value=note)

# v3.5p shadow columns appended (Class v3.5p, Mult v3.5p, YA Flag, YA Reason)
cls_cols = ["Pitcher", "Classification", "Class Reason", "Multiplier", "SwStr%", "Whiff%",
            "Class v3.5p (shadow)", "Mult v3.5p", "YA Flag", "YA Reason"]
# Shadow disclaimer just under tab title
ws["B10"] = "v3.5p columns are SHADOW only — BF gate is unchanged today. Compare via backtest after 7 days."
ws["B10"].font = Font(name="Calibri", italic=True, size=9, color="595959")
ws.merge_cells("B10:K10")
style_header(ws, 11, cls_cols)
ws.freeze_panes = "A12"
r = 12
for row in results:
    vals = [row["name"], row.get("Classification", ""), row.get("Class Reason", ""),
            row.get("Multiplier"), row.get("SwStr%"), row.get("Whiff%"),
            row.get("Class v3.5p"), row.get("Mult v3.5p"),
            row.get("v3.5p Young Arm Flag"), row.get("v3.5p Young Arm Reason")]
    for ci, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.border = THIN_BORDER
        if ci == 5 and isinstance(v, (int, float)):
            cell.number_format = "0.00"
        elif ci > 5 and ci <= 7 and isinstance(v, (int, float)):
            cell.number_format = "0.00"
        elif ci == 9 and isinstance(v, (int, float)):
            cell.number_format = "0.00"
        # Highlight shadow disagreement with v3.5
        if ci == 8 and v and v != row.get("Classification"):
            cell.fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow
        if ci == 10 and v == "BLOCK":
            cell.fill = PatternFill("solid", fgColor="FCE4D6")  # light salmon
    r += 1
for ci, w in enumerate([22, 16, 30, 12, 10, 10, 24, 10, 8, 55]):
    ws.column_dimensions[get_column_letter(2 + ci)].width = w


# ── TAB 4: Lines & Movement ────────────────────────────────────────────────
ws = wb.create_sheet("Lines & Movement")
ws.column_dimensions["A"].width = 3
ws["B2"] = "Lines & Movement — DraftKings + FanDuel (The Odds API)"
ws["B2"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
ws.merge_cells("B2:M2")
ws["B3"] = "Note: This is the 11 AM run — line movement vs late refresh will populate in the *_late workbook."
ws["B3"].font = Font(name="Calibri", italic=True, size=9, color="595959")
ws.merge_cells("B3:M3")

ln_cols = ["Pitcher", "Matchup", "DK Line", "DK Over", "DK Under", "FD Line", "FD Over", "FD Under",
           "Best Line", "Best Under", "Best Under Book", "Shop Under", "Shop Under Book"]
style_header(ws, 4, ln_cols)
ws.freeze_panes = "A5"
r = 5
for row in results:
    vals = [row["name"], row.get("matchup", ""),
            row.get("DK Line"), row.get("DK Over"), row.get("DK Under"),
            row.get("FD Line"), row.get("FD Over"), row.get("FD Under"),
            row.get("Best Line"), row.get("Best Under"), row.get("Best Under Book"),
            row.get("Shop Under"), row.get("Shop Under Book")]
    for ci, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        # Color the Best Under cell if it meets the -130 gate
        if ci == 11 and isinstance(v, (int, float)) and v <= -130:
            cell.fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
        if ci in (4, 5, 7, 8) and isinstance(v, (int, float)):
            cell.number_format = "+0;-0"
        elif ci in (3, 6, 9) and isinstance(v, (int, float)):
            cell.number_format = "0.0"
        elif ci in (10, 12) and isinstance(v, (int, float)):
            cell.number_format = "+0;-0"
    r += 1
for ci, w in enumerate([22, 14, 8, 8, 8, 8, 8, 8, 9, 11, 14, 11, 14]):
    ws.column_dimensions[get_column_letter(2 + ci)].width = w


# ── TAB 5: Weather & Park ──────────────────────────────────────────────────
ws = wb.create_sheet("Weather & Park")
ws.column_dimensions["A"].width = 3
ws["B2"] = "Weather, Park, and Opposing Lineup"
ws["B2"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
ws.merge_cells("B2:K2")
ws["B3"] = "Park factor source: v3.6.2 (Z-Files 2026, as of 2026-05-29) — lineup-weighted by R/L/S handedness counts."
ws["B3"].font = Font(name="Calibri", italic=True, size=9, color="595959")
ws.merge_cells("B3:K3")

wx_cols = ["Pitcher", "Matchup", "First Pitch", "Park", "Park K Shift (overall)",
           "Opp Lineup (R/L/S)", "Weather", "Lineup Confirmed", "PPD%"]
style_header(ws, 4, wx_cols)
ws.freeze_panes = "A5"
r = 5
for row in results:
    ppd_raw = row.get("PPD%", 0) or 0
    try:
        ppd = float(str(ppd_raw).replace('%','').strip() or 0)
    except Exception:
        ppd = 0
    ppd_fill = PatternFill(start_color=PINK_FILL, end_color=PINK_FILL, fill_type="solid") if ppd > 50 else None
    vals = [row["name"], row.get("matchup", ""), row.get("time", ""),
            row.get("park", ""), row.get("park_factor"),
            row.get("Opp Lineup", ""), row.get("Weather", ""),
            row.get("Lineup Confirmed", ""), ppd]
    for ci, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.border = THIN_BORDER
        cell.alignment = CENTER if ci != 8 else LEFT
        if ci == 6 and isinstance(v, (int, float)):
            cell.number_format = "+0.00;-0.00;0.00"
        if ci == 10 and isinstance(v, (int, float)):
            cell.number_format = "0%"
            if ppd_fill:
                cell.fill = ppd_fill
    r += 1
for ci, w in enumerate([22, 14, 12, 22, 12, 14, 26, 13, 8]):
    ws.column_dimensions[get_column_letter(2 + ci)].width = w


# ── TAB 6: Model Run ────────────────────────────────────────────────────────
ws = wb.create_sheet("Model Run")
ws.column_dimensions["A"].width = 3
ws["B2"] = "Model Run — v3.5 (production) + v4.0 SHADOW"
ws["B2"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
ws.merge_cells("B2:P2")
ws["B3"] = "v3.5 = (K/9 × IP/GS / 9) × Multiplier + env adj → v3.3 4-stage calibration. v4.0 = K/9 × projected_IP / 9 + same env → v3.3 cal. Both with v3.6.2 park factors."
ws["B3"].font = Font(name="Calibri", italic=True, size=9, color="595959")
ws.merge_cells("B3:P3")

mr_cols = ["Pitcher", "Matchup", "Park", "Park Shift", "Opp Lu", "Hand",
           "Raw v3.5", "Proj Ks (v3.5)", "Proj Ks v4.0", "Δ (v4-v3.5)",
           "Best Line", "v3.5 Resid", "v4.0 Resid", "Rule 5", "Rule 5 Reason"]
style_header(ws, 4, mr_cols)
ws.freeze_panes = "A5"
r = 5
for row in results:
    is_watch = row.get("BF Call") == "Watch List"
    row_fill = PatternFill(start_color=GRAY_FILL, end_color=GRAY_FILL, fill_type="solid") if is_watch else None
    delta = row.get("Delta v4-v35")
    delta_fill = None
    if isinstance(delta, (int, float)):
        if delta > 0.05:
            delta_fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
        elif delta < -0.05:
            delta_fill = PatternFill(start_color=RED_FILL, end_color=RED_FILL, fill_type="solid")
    vals = [
        row["name"], row.get("matchup", ""), row.get("park", ""),
        row.get("park_factor"), row.get("Opp Lineup", ""), row.get("hand", ""),
        row.get("Raw Ks"), row.get("Proj Ks"), row.get("Proj Ks v4.0"),
        row.get("Delta v4-v35"), row.get("Best Line"),
        row.get("Residual"), row.get("Residual v4.0"),
        row.get("Rule 5", ""), row.get("Rule 5 Reason", ""),
    ]
    for ci, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.border = THIN_BORDER
        cell.alignment = CENTER if ci not in (2, 16) else LEFT
        if ci == 5 and isinstance(v, (int, float)):
            cell.number_format = "+0.00;-0.00;0.00"
        elif ci in (8, 9, 10) and isinstance(v, (int, float)):
            cell.number_format = "0.00"
        elif ci == 11 and isinstance(v, (int, float)):
            cell.number_format = "+0.00;-0.00;0.00"
        elif ci == 12 and isinstance(v, (int, float)):
            cell.number_format = "0.0"
        elif ci in (13, 14) and isinstance(v, (int, float)):
            cell.number_format = "+0.00;-0.00;0.00"
        if ci == 11 and delta_fill:
            cell.fill = delta_fill
        elif row_fill:
            cell.fill = row_fill
    r += 1
for ci, w in enumerate([22, 14, 22, 10, 11, 7, 10, 13, 13, 12, 10, 11, 11, 9, 28]):
    ws.column_dimensions[get_column_letter(2 + ci)].width = w


# ── TAB 7: BF Picks ────────────────────────────────────────────────────────
ws = wb.create_sheet("BF Picks")
ws.column_dimensions["A"].width = 3
ws["B2"] = "BF Picks — v3.5 GATE (residual ≤ -0.75, best Under ≤ -130, Rule 5 pass, PPD ≤ 50%)"
ws["B2"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
ws.merge_cells("B2:N2")
ws["B3"] = "BF gates remain on v3.5 during 7-day shadow period for v4.0."
ws["B3"].font = Font(name="Calibri", italic=True, size=9, color="595959")
ws.merge_cells("B3:N3")

bf_cols = ["Pitcher", "Team", "Matchup", "First Pitch", "Best Line", "Best Under (gate)",
           "Shop Under", "Proj Ks (v3.5)", "Proj Ks v4.0", "v3.5 Resid", "v4.0 Resid",
           "PPD%", "Direction", "BF Reason"]
style_header(ws, 4, bf_cols)
ws.freeze_panes = "A5"
r = 5
bf_rows = [row for row in results if row.get("BF Call") == "Under"]
if not bf_rows:
    ws.cell(row=5, column=2, value="No BF picks today.").font = Font(italic=True, color="595959")
    ws.cell(row=5, column=2).alignment = LEFT
else:
    for row in bf_rows:
        vals = [row["name"], row.get("team", ""), row.get("matchup", ""), row.get("time", ""),
                row.get("Best Line"), f"{row.get('Best Under')} ({row.get('Best Under Book')})",
                f"{row.get('Shop Under')} ({row.get('Shop Under Book')})",
                row.get("Proj Ks"), row.get("Proj Ks v4.0"),
                row.get("Residual"), row.get("Residual v4.0"),
                row.get("PPD%"), "Under", row.get("BF Reason", "")]
        for ci, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = THIN_BORDER
            cell.fill = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")
            cell.alignment = CENTER if ci not in (2, 4, 15) else LEFT
            if ci == 6 and isinstance(v, (int, float)):
                cell.number_format = "0.0"
            elif ci in (9, 10) and isinstance(v, (int, float)):
                cell.number_format = "0.00"
            elif ci in (11, 12) and isinstance(v, (int, float)):
                cell.number_format = "+0.00;-0.00;0.00"
            elif ci == 13 and isinstance(v, (int, float)):
                cell.number_format = "0%"
        r += 1
for ci, w in enumerate([22, 7, 14, 12, 10, 18, 18, 13, 13, 11, 11, 8, 10, 50]):
    ws.column_dimensions[get_column_letter(2 + ci)].width = w


# ── TAB 8: Watch List ──────────────────────────────────────────────────────
ws = wb.create_sheet("Watch List")
ws.column_dimensions["A"].width = 3
ws["B2"] = "Watch List — Rule 5 fails, FG cohort misses, or non-priced starters"
ws["B2"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
ws.merge_cells("B2:I2")

wl_cols = ["Pitcher", "Team", "Matchup", "First Pitch", "GS", "IP", "K/9", "FG Source", "Reason"]
style_header(ws, 3, wl_cols)
ws.freeze_panes = "A4"
r = 4
for row in results:
    if row.get("BF Call") != "Watch List":
        continue
    vals = [row["name"], row.get("team", ""), row.get("matchup", ""), row.get("time", ""),
            row.get("GS"), row.get("IP"), row.get("K/9"),
            row.get("FG Source", ""), row.get("BF Reason", "") or row.get("Rule 5 Reason", "")]
    for ci, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color=GRAY_FILL, end_color=GRAY_FILL, fill_type="solid")
        cell.alignment = CENTER if ci not in (2, 4, 10) else LEFT
        if ci == 8 and isinstance(v, (int, float)):
            cell.number_format = "0.00"
    r += 1
for ci, w in enumerate([22, 7, 14, 12, 6, 8, 7, 12, 36]):
    ws.column_dimensions[get_column_letter(2 + ci)].width = w


# Save
OUT = str(WORKSPACE / f"MLB_K_Prop_Model_{DATE_STR}_{SLATE_TYPE}.xlsx")
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Sheets: {wb.sheetnames}")
